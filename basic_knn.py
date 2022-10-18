import openpyxl
import math

# this is the most basic KNN I think possible, but now I know where I'm heading from here
# time to learn how to apply it with a combination of other tools to make more complex predictions :)

class testingPoint:
    def __init__(self, x, y):
        self.x = x
        self.y = y

class trainingPoint:
    def __init__(self, x, y, info):
        self.x = x
        self.y = y
        self.info = info

def get_training_data(training_filepath):
    wb = openpyxl.load_workbook(training_filepath)
    sheet = wb.active
    max_row = sheet.max_row
    max_col = sheet.max_column
    all_rows = []
    for x in range(max_row):
        this_row = []
        for y in range(max_col):
            this_row.append(sheet.cell(row=x+1, column=y+1).value)
        all_rows.append(trainingPoint(this_row[0], this_row[1], this_row[2]))
    return all_rows

def get_testing_data(testing_filepath):
    wb = openpyxl.load_workbook(testing_filepath)
    sheet = wb.active
    max_row = sheet.max_row
    max_col = sheet.max_column
    all_rows = []
    for x in range(max_row):
        this_row = []
        for y in range(max_col):
            this_row.append(sheet.cell(row=x + 1, column=y + 1).value)
        all_rows.append(testingPoint(this_row[0], this_row[1]))
    return all_rows

def point_dist(point_a, point_b):
    # point format in [x,y]
    return math.sqrt(((abs(point_a[0] - point_b[0]))**2) + ((abs(point_a[1] - point_b[1]))**2))

def run_predictions(testing_data, training_data, k):
    all_predictions = []
    for x in range(len(testing_data)):
        test = testing_data[x]
        all_dists = {}
        for train in training_data:
            this_dist = point_dist([test.x, test.y], [train.x, train.y])
            all_dists[train] = this_dist
        sorted_dists = sorted(all_dists.items(), key=lambda kv: kv[1])
        k_dists = sorted_dists[0:k]
        freq_dict = {}
        for dist_info in k_dists:
            this_dist_info = dist_info[0].info
            if this_dist_info in freq_dict:
                freq_dict[this_dist_info] += 1
            else:
                freq_dict[this_dist_info] = 1
        all_predictions.append([[test.x, test.y], max(freq_dict, key=freq_dict.get)])
    return all_predictions








